# -*- coding: utf-8 -*-
"""
File: fill_csv.py
Purpose: Automatically match ore serial numbers and fill assay grades from CSV into Excel sheet.
Author: Antigravity
Date: 2026-05-29
"""

import os
import pandas as pd
import openpyxl

def fill_assay_grades(excel_path: str, csv_path: str, sheet_name: str = '0514氧化铜') -> int:
    """
    将 CSV 文件中的化验品位（Cu, Fe, Al, Ca, S）匹配并填入 Excel 工作簿中指定 Sheet 的正面化验品位列中，
    同时将 CSV 文件的第二列（测试 # / XRF测试序号）填入 Excel 的“XRF编号”列（C列，第 3 列）中。

    参数说明 (Parameters):
    ---------------------
    excel_path : str
        类型：字符串。
        含义：目标 Excel 文件的绝对或相对路径（例如 'fill_csv/CuO矿石重量.xlsx'）。
        用法：该文件将被 openpyxl 加载、原地更新指定 Sheet 单元格后保存。该操作保留原 Excel 的所有格式、公式和其他工作表。
    
    csv_path : str
        类型：字符串.
        含义：包含源化验品位数据的 CSV 文件路径（例如 'fill_csv/2026.05.29.csv'）。
        用法：使用 pandas 读取该文件。第二列（索引为 1）为 XRF 测试序号，第三列（索引为 2）为矿石序号，数据中应包含 'Cu', 'Fe', 'Al', 'Ca', 'S' 列作为品位数据。
    
    sheet_name : str
        类型：字符串，默认值为 '0514氧化铜'。
        含义：要更新的 Excel 工作表名称。
        用法：指定写入数据的目标 Sheet 名称。

    返回值 (Returns):
    -----------------
    int
        含义：成功匹配并填入的记录条数。

    用法示例 (Usage Example):
    -------------------------
    >>> fill_assay_grades('fill_csv/CuO矿石重量.xlsx', 'fill_csv/2026.05.29.csv')
    """
    print(f"[*] 开始读取数据文件...")
    print(f"    Excel 路径: {excel_path}")
    print(f"    CSV 路径  : {csv_path}")
    print(f"    目标 Sheet: {sheet_name}")

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"未找到 Excel 文件: {excel_path}")
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"未找到 CSV 文件: {csv_path}")

    # 1. 使用 pandas 读取 CSV 校验数据
    # 指定 gbk 编码解决中文读取问题
    df_csv = pd.read_csv(csv_path, encoding='gbk')
    
    # 获取第二列为测试序号列，第三列为矿石序号列
    test_col = df_csv.columns[1]
    id_col = df_csv.columns[2]
    print(f"[*] 成功加载 CSV 数据，识别到测试序号列为: '{test_col}'，矿石序号列为: '{id_col}'")
    
    # 提取目标元素列表
    target_elements = ['Cu', 'Fe', 'Al', 'Ca', 'S']
    for el in target_elements:
        if el not in df_csv.columns:
            raise KeyError(f"CSV 文件中缺少目标元素列: '{el}'")

    # 将 CSV 数据整理为以序号为键的字典，便于 O(1) 查找匹配
    csv_data = {}
    for idx, row in df_csv.iterrows():
        try:
            ore_id = int(float(row[id_col]))
        except (ValueError, TypeError):
            print(f"[!] CSV 第 {idx} 行的序号无法解析为整数: {row[id_col]}，跳过该行。")
            continue
        
        # 提取第二列的 XRF 测试序号并转为整数或字符串
        try:
            test_no = int(float(row[test_col]))
        except (ValueError, TypeError):
            test_no = str(row[test_col])
        
        # 整理该序号对应的元素品位，如果为 NaN 则处理为 None 以在 Excel 中留空
        grades = {}
        for el in target_elements:
            val = row[el]
            grades[el] = None if pd.isna(val) else float(val)
        
        csv_data[ore_id] = {
            'grades': grades,
            'test_no': test_no
        }

    print(f"[*] 从 CSV 中提取了 {len(csv_data)} 条有效矿石品位与测试序号记录，准备填入 Excel。")

    # 2. 使用 openpyxl 加载 Excel（data_only=False 确保公式不被破坏）
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 工作簿中未找到名为 '{sheet_name}' 的 Sheet。可选的 Sheet 有: {wb.sheetnames}")
    
    ws = wb[sheet_name]

    # 定义目标元素写入的列索引 (1-based)
    # 根据分析：
    # Column A (1) -> 序号
    # Column C (3) -> XRF测试序号 / XRF编号
    # Column D (4) -> 正面 Cu
    # Column E (5) -> 正面 Fe
    # Column F (6) -> 正面 Al
    # Column G (7) -> 正面 Ca
    # Column H (8) -> 正面 S
    element_to_col = {
        'Cu': 4,
        'Fe': 5,
        'Al': 6,
        'Ca': 7,
        'S': 8
    }

    match_count = 0
    # 遍历 Excel 工作表的每一行（从第 3 行开始，避开表头）
    for r in range(3, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is None:
            continue
            
        try:
            excel_id = int(float(cell_val))
        except (ValueError, TypeError):
            # 无法解析为序号整数，说明是非数据行，忽略
            continue
        
        # 如果该序号在 CSV 数据中存在，则填入品位和 XRF 测试序号
        if excel_id in csv_data:
            entry = csv_data[excel_id]
            grades = entry['grades']
            test_no = entry['test_no']
            
            # 写入正面品位
            for el, col_idx in element_to_col.items():
                val = grades[el]
                ws.cell(row=r, column=col_idx, value=val)
            
            # 写入 XRF测试序号 (C列, 第3列)
            ws.cell(row=r, column=3, value=test_no)
            
            print(f"    [+] 成功更新 Excel 序号 {excel_id} (行号 {r}) 的正面品位: {grades}，XRF编号: {test_no}")
            match_count += 1

    # 3. 保存更新后的工作簿
    wb.save(excel_path)
    wb.close()
    
    print(f"[*] 数据填充完毕！共成功匹配并写入 {match_count} 条记录。Excel 文件已妥善保存。")
    return match_count

if __name__ == '__main__':
    # 获取脚本文件所在的目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 拼接出 Excel 和 CSV 的绝对路径，保证脚本在任何目录下运行都有效
    default_excel_path = os.path.join(script_dir, 'CuO矿石重量.xlsx')
    default_csv_path = os.path.join(script_dir, '2026.05.29.csv')
    
    # 执行填充
    try:
        updated_rows = fill_assay_grades(default_excel_path, default_csv_path)
        print(f"\n[Success] 任务圆满完成，更新了 {updated_rows} 行数据。")
    except Exception as e:
        print(f"\n[Error] 执行过程中发生错误: {e}")
